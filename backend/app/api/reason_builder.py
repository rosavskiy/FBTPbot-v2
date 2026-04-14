"""
API «Помощник формирования причин обращений».

Два основных блока:
1. Анализ пересечений маркеров между существующими причинами + LLM-рекомендации.
2. Генерация причин обращений из загруженной таблицы (CSV/XLSX) через LLM.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
import uuid

import httpx
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.api.admin_auth import require_role, verify_admin_token
from app.database.reason_store import get_all_reasons, upsert_reason
from app.llm_settings import get_llm_settings_snapshot
from app.models.reason_schemas import ContactReason

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/reason-builder", tags=["reason-builder"])

_editor = Depends(require_role("superadmin", "admin"))
_any_admin = Depends(verify_admin_token)

# Temporary storage for uploaded tables (session-based)
_uploaded_tables: dict[str, dict] = {}

# ── Request / Response models ──


class OverlapItem(BaseModel):
    reason_a_id: str
    reason_a_name: str
    reason_b_id: str
    reason_b_name: str
    overlap_type: str
    shared_markers: list[str]
    severity: int  # phrase=40, numeric=30, noun=20, verb=10


class OverlapsResponse(BaseModel):
    total_reasons: int
    total_overlaps: int
    overlaps: list[OverlapItem]


class RecommendRequest(BaseModel):
    overlaps: list[OverlapItem]
    llm_provider: str = "yandex"


class RecommendResponse(BaseModel):
    recommendations: str


class UploadedReason(BaseModel):
    name: str
    ticket_count: int
    sample_texts: list[str]


class UploadResponse(BaseModel):
    session_id: str
    total_tickets: int
    reasons: list[UploadedReason]


class GenerateRequest(BaseModel):
    session_id: str
    reason_name: str
    llm_provider: str = "yandex"


class GenerateBatchRequest(BaseModel):
    session_id: str
    reason_names: list[str]
    llm_provider: str = "yandex"


class GeneratedReason(BaseModel):
    id: str
    name: str
    markers: dict
    ticket_count: int = 0


class GenerateBatchResponse(BaseModel):
    results: list[GeneratedReason]
    errors: list[dict] = []


class AddReasonRequest(BaseModel):
    reason: dict


class ExportReasonRequest(BaseModel):
    reason: dict


# ── Severity weights ──
_SEVERITY = {"phrase_masks": 40, "numeric_tags": 30, "nouns": 20, "verbs": 10}

# ── LLM prompts ──

_GENERATE_MARKERS_PROMPT = """Ты помогаешь создавать базу знаний для ИИ-бота техподдержки аптечного ПО.

Проанализируй список заявок по причине обращения «{reason_name}» и выдели:

1. **Глаголы-маркеры** — действия пользователя (в инфинитиве). Важно: только глаголы, реально описывающие суть проблемы.
2. **Существительные-маркеры** — объекты, с которыми работает пользователь. Только термины ПО, документы, сущности.
3. **Числовые теги** — коды ошибок, номера форм и т.п. Только если одно число повторяется в нескольких заявках.
4. **Фразовые маски** — точные фразы (2+ слов) из текста заявок, однозначно указывающие на эту причину. Должны быть максимально специфичными, чтобы не пересекаться с другими причинами.

ВАЖНЫЕ ПРАВИЛА ФИЛЬТРАЦИИ:
- Если текст заявки содержит лишние вопросы (\"как дела?\", \"почему не отвечаете?\") — ИГНОРИРУЙ их.
- Если текст смутный или бессмысленный — бери только понятную суть.
- Если пользователь пишет несколько проблем — бери только ту, что соответствует причине.
- Если текст заявки — это просто \"здравствуйте\" или \"помогите\" — не делай из него маркеры.

КОНТЕКСТ ВЕСОВ:
- Фразовая маска: 10.0 (самый высокий вес — должна быть уникальной для этой причины)
- Числовой тег: 5.0
- Существительное: 2.0
- Глагол: 1.0

Ответь СТРОГО в формате JSON (без markdown, без пояснений):
{{"verbs": ["глагол1", "глагол2"], "nouns": ["сущ1", "сущ2"], "numeric_tags": ["11", "54"], "phrase_masks": ["точная фраза 1", "точная фраза 2"]}}

ЗАЯВКИ ПО ПРИЧИНЕ «{reason_name}»:
{texts}
"""

_RECOMMEND_OVERLAPS_PROMPT = """Ты — эксперт по настройке ИИ-бота техподдержки.

Есть система классификации заявок по причинам обращения. Каждая причина имеет маркеры:
- Фразовые маски (вес 10.0) — самые точные, не должны пересекаться
- Числовые теги (вес 5.0)
- Существительные (вес 2.0)
- Глаголы (вес 1.0)

Обнаружены ПЕРЕСЕЧЕНИЯ маркеров между причинами, что может приводить к НЕПРАВИЛЬНОЙ классификации заявок.

Для каждого пересечения ниже дай КОНКРЕТНУЮ рекомендацию:
- Какие маркеры убрать и из какой причины
- Какие фразовые маски переформулировать (сделать уникальнее)
- Общие советы по разграничению причин

ПЕРЕСЕЧЕНИЯ:
{overlaps_text}

Формат ответа: по каждому пересечению — номер, причины, рекомендация. Пиши на русском, кратко и по делу.
"""


# ── Helpers ──


def _slugify(text: str) -> str:
    """Convert name to ID: lowercase, transliterate, underscores."""
    tr = {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "е": "e",
        "ё": "yo",
        "ж": "zh",
        "з": "z",
        "и": "i",
        "й": "j",
        "к": "k",
        "л": "l",
        "м": "m",
        "н": "n",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "у": "u",
        "ф": "f",
        "х": "kh",
        "ц": "ts",
        "ч": "ch",
        "ш": "sh",
        "щ": "sch",
        "ъ": "",
        "ы": "y",
        "ь": "",
        "э": "e",
        "ю": "yu",
        "я": "ya",
    }
    result = []
    for ch in text.lower():
        if ch in tr:
            result.append(tr[ch])
        elif ch.isalnum():
            result.append(ch)
        else:
            result.append("_")
    slug = re.sub(r"_+", "_", "".join(result)).strip("_")
    return slug[:64] if slug else "reason"


def _normalize_provider(provider: str) -> str:
    normalized = (provider or "").strip().lower()
    if normalized not in {"yandex", "deepseek"}:
        raise HTTPException(status_code=422, detail="Допустимые провайдеры: yandex, deepseek")
    return normalized


async def _llm_complete(provider: str, system_prompt: str, user_prompt: str, temperature: float = 0.1) -> str:
    """Call LLM via configured provider (reuse keys from llm_settings)."""
    provider = _normalize_provider(provider)
    llm_settings = get_llm_settings_snapshot()

    if provider == "deepseek":
        url = "https://api.deepseek.com/chat/completions"
        headers = {
            "Authorization": f"Bearer {llm_settings['deepseek_api_key']}",
            "Content-Type": "application/json",
        }
        body = {
            "model": llm_settings.get("deepseek_model", "deepseek-chat"),
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "temperature": temperature,
            "max_tokens": 4000,
            "stream": False,
        }
        async with httpx.AsyncClient(timeout=120.0) as client:
            resp = await client.post(url, headers=headers, json=body)
        if resp.status_code != 200:
            logger.error(f"DeepSeek API error {resp.status_code}: {resp.text}")
            raise HTTPException(status_code=502, detail=f"LLM API error: {resp.status_code}")
        return resp.json()["choices"][0]["message"]["content"]
    else:
        # Yandex GPT
        url = "https://llm.api.cloud.yandex.net/foundationModels/v1/completion"
        headers = {
            "Authorization": f"Api-Key {llm_settings['yandex_api_key']}",
            "Content-Type": "application/json",
        }
        body = {
            "modelUri": f"gpt://{llm_settings['yandex_folder_id']}/{llm_settings.get('yandex_gpt_model', 'yandexgpt')}/latest",
            "completionOptions": {
                "stream": False,
                "temperature": temperature,
                "maxTokens": "4000",
            },
            "messages": [
                {"role": "system", "text": system_prompt},
                {"role": "user", "text": user_prompt},
            ],
        }
        async with httpx.AsyncClient(timeout=120.0) as client:
            resp = await client.post(url, headers=headers, json=body)
        if resp.status_code != 200:
            logger.error(f"YandexGPT API error {resp.status_code}: {resp.text}")
            raise HTTPException(status_code=502, detail=f"LLM API error: {resp.status_code}")
        return resp.json()["result"]["alternatives"][0]["message"]["text"]


def _detect_csv_params(raw: bytes) -> tuple[str, str]:
    """Detect encoding and delimiter for CSV."""
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            text = raw.decode(encoding)
            # Check for common delimiters
            first_lines = text.split("\n", 5)[:5]
            for delim in (";", ",", "\t"):
                if all(delim in line for line in first_lines if line.strip()):
                    return encoding, delim
            return encoding, ";"
        except (UnicodeDecodeError, ValueError):
            continue
    return "utf-8", ";"


def _parse_csv(raw: bytes) -> list[dict]:
    """Parse CSV into list of {reason, text} dicts."""
    encoding, delimiter = _detect_csv_params(raw)
    text = raw.decode(encoding)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        raise HTTPException(status_code=422, detail="Файл пуст")

    # Skip header if it looks like one
    header = rows[0]
    start = 0
    if len(header) >= 2:
        h0 = header[0].strip().lower()
        if any(kw in h0 for kw in ("причин", "reason", "тема", "категори", "№", "#", "номер")):
            start = 1

    result = []
    for row in rows[start:]:
        if len(row) < 2:
            continue
        reason_name = row[0].strip()
        ticket_text = row[1].strip()
        if reason_name and ticket_text:
            result.append({"reason": reason_name, "text": ticket_text})
    return result


def _parse_xlsx(raw: bytes) -> list[dict]:
    """Parse XLSX into list of {reason, text} dicts."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl не установлен. Добавьте в requirements.txt.")

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=422, detail="XLSX файл не содержит листов")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=422, detail="Файл пуст")

    # Skip header
    start = 0
    if rows[0] and len(rows[0]) >= 2:
        h0 = str(rows[0][0] or "").strip().lower()
        if any(kw in h0 for kw in ("причин", "reason", "тема", "категори", "№", "#", "номер")):
            start = 1

    result = []
    for row in rows[start:]:
        if not row or len(row) < 2:
            continue
        reason_name = str(row[0] or "").strip()
        ticket_text = str(row[1] or "").strip()
        if reason_name and ticket_text:
            result.append({"reason": reason_name, "text": ticket_text})

    wb.close()
    return result


def _group_by_reason(tickets: list[dict]) -> dict[str, list[str]]:
    """Group tickets by reason name."""
    groups: dict[str, list[str]] = {}
    for t in tickets:
        name = t["reason"]
        groups.setdefault(name, []).append(t["text"])
    return groups


def _find_overlaps(reasons: list[ContactReason]) -> list[OverlapItem]:
    """Find marker overlaps between all active reason pairs."""
    overlaps: list[OverlapItem] = []

    for i, a in enumerate(reasons):
        for b in reasons[i + 1 :]:
            # Phrase masks — exact match
            shared_phrases = {p.lower().strip() for p in a.markers.phrase_masks} & {
                p.lower().strip() for p in b.markers.phrase_masks
            }
            if shared_phrases:
                overlaps.append(
                    OverlapItem(
                        reason_a_id=a.id,
                        reason_a_name=a.name,
                        reason_b_id=b.id,
                        reason_b_name=b.name,
                        overlap_type="phrase_masks",
                        shared_markers=sorted(shared_phrases),
                        severity=_SEVERITY["phrase_masks"],
                    )
                )

            # Numeric tags
            shared_numeric = set(a.markers.numeric_tags) & set(b.markers.numeric_tags)
            if shared_numeric:
                overlaps.append(
                    OverlapItem(
                        reason_a_id=a.id,
                        reason_a_name=a.name,
                        reason_b_id=b.id,
                        reason_b_name=b.name,
                        overlap_type="numeric_tags",
                        shared_markers=sorted(shared_numeric),
                        severity=_SEVERITY["numeric_tags"],
                    )
                )

            # Nouns — lemma-based
            shared_nouns = {n.lower().strip() for n in a.markers.nouns} & {n.lower().strip() for n in b.markers.nouns}
            if shared_nouns:
                overlaps.append(
                    OverlapItem(
                        reason_a_id=a.id,
                        reason_a_name=a.name,
                        reason_b_id=b.id,
                        reason_b_name=b.name,
                        overlap_type="nouns",
                        shared_markers=sorted(shared_nouns),
                        severity=_SEVERITY["nouns"],
                    )
                )

            # Verbs
            shared_verbs = {v.lower().strip() for v in a.markers.verbs} & {v.lower().strip() for v in b.markers.verbs}
            if shared_verbs:
                overlaps.append(
                    OverlapItem(
                        reason_a_id=a.id,
                        reason_a_name=a.name,
                        reason_b_id=b.id,
                        reason_b_name=b.name,
                        overlap_type="verbs",
                        shared_markers=sorted(shared_verbs),
                        severity=_SEVERITY["verbs"],
                    )
                )

    overlaps.sort(key=lambda o: (-o.severity, o.reason_a_name))
    return overlaps


def _extract_json_from_llm(text: str) -> dict:
    """Extract JSON object from LLM response (handles markdown fences)."""
    # Try to find JSON in code blocks
    code_match = re.search(r"```(?:json)?\s*\n?(.*?)\n?```", text, re.DOTALL)
    if code_match:
        text = code_match.group(1)

    # Find first { ... }
    brace_start = text.find("{")
    if brace_start == -1:
        raise ValueError("No JSON object found in LLM response")

    depth = 0
    for i in range(brace_start, len(text)):
        if text[i] == "{":
            depth += 1
        elif text[i] == "}":
            depth -= 1
            if depth == 0:
                try:
                    return json.loads(text[brace_start : i + 1])
                except json.JSONDecodeError:
                    raise ValueError("Invalid JSON in LLM response")

    raise ValueError("Unclosed JSON object in LLM response")


# ── Endpoints ──


@router.post("/analyze-overlaps", response_model=OverlapsResponse, dependencies=[_any_admin])
async def analyze_overlaps():
    """Программный анализ пересечений маркеров между всеми активными причинами."""
    reasons = get_all_reasons(active_only=False)
    overlaps = _find_overlaps(reasons)
    return OverlapsResponse(
        total_reasons=len(reasons),
        total_overlaps=len(overlaps),
        overlaps=overlaps,
    )


@router.post("/analyze-overlaps/recommend", response_model=RecommendResponse, dependencies=[_editor])
async def recommend_overlaps(req: RecommendRequest):
    """LLM-рекомендации по найденным пересечениям маркеров."""
    if not req.overlaps:
        return RecommendResponse(recommendations="Пересечений не найдено.")

    # Limit to top 30 overlaps for context size
    top_overlaps = req.overlaps[:30]

    type_labels = {
        "phrase_masks": "Фразовые маски (вес 10.0)",
        "numeric_tags": "Числовые теги (вес 5.0)",
        "nouns": "Существительные (вес 2.0)",
        "verbs": "Глаголы (вес 1.0)",
    }

    lines = []
    for i, o in enumerate(top_overlaps, 1):
        label = type_labels.get(o.overlap_type, o.overlap_type)
        joined = ", ".join(o.shared_markers[:10])
        lines.append(f"{i}. [{label}] «{o.reason_a_name}» ↔ «{o.reason_b_name}»: {joined}")

    overlaps_text = "\n".join(lines)
    provider = _normalize_provider(req.llm_provider)

    try:
        result = await _llm_complete(
            provider,
            "Ты — эксперт по настройке классификации заявок техподдержки.",
            _RECOMMEND_OVERLAPS_PROMPT.format(overlaps_text=overlaps_text),
            temperature=0.2,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"LLM recommendation error: {e}")
        raise HTTPException(status_code=502, detail=f"Ошибка LLM: {e}")

    return RecommendResponse(recommendations=result)


@router.post("/upload-table", response_model=UploadResponse, dependencies=[_editor])
async def upload_table(file: UploadFile = File(...)):
    """Загрузить CSV/XLSX таблицу с причинами обращений и текстами заявок."""
    if not file.filename:
        raise HTTPException(status_code=422, detail="Имя файла не указано")

    raw = await file.read()
    if len(raw) > 50 * 1024 * 1024:
        raise HTTPException(status_code=422, detail="Файл слишком большой (макс. 50 МБ)")

    ext = (file.filename.rsplit(".", 1)[-1] if "." in file.filename else "").lower()

    if ext == "xlsx":
        tickets = _parse_xlsx(raw)
    elif ext in ("csv", "txt", "tsv"):
        tickets = _parse_csv(raw)
    else:
        raise HTTPException(status_code=422, detail="Формат не поддерживается. Используйте CSV или XLSX.")

    if not tickets:
        raise HTTPException(status_code=422, detail="Не удалось извлечь данные из файла.")

    groups = _group_by_reason(tickets)

    session_id = str(uuid.uuid4())
    _uploaded_tables[session_id] = groups

    # Cleanup old sessions (keep max 20)
    if len(_uploaded_tables) > 20:
        oldest_keys = list(_uploaded_tables.keys())[:-20]
        for k in oldest_keys:
            _uploaded_tables.pop(k, None)

    reasons_list = []
    for name, texts in sorted(groups.items(), key=lambda x: -len(x[1])):
        reasons_list.append(
            UploadedReason(
                name=name,
                ticket_count=len(texts),
                sample_texts=texts[:5],
            )
        )

    return UploadResponse(
        session_id=session_id,
        total_tickets=len(tickets),
        reasons=reasons_list,
    )


@router.post("/generate-reason", dependencies=[_editor])
async def generate_reason(req: GenerateRequest):
    """Сгенерировать маркеры для одной причины через LLM."""
    groups = _uploaded_tables.get(req.session_id)
    if groups is None:
        raise HTTPException(status_code=404, detail="Сессия не найдена. Загрузите таблицу заново.")

    texts = groups.get(req.reason_name)
    if texts is None:
        raise HTTPException(status_code=404, detail=f"Причина «{req.reason_name}» не найдена в загруженных данных.")

    provider = _normalize_provider(req.llm_provider)

    # Limit texts to avoid token overflow (max ~100 texts, each truncated)
    limited_texts = [t[:500] for t in texts[:150]]
    texts_block = "\n".join(f"- {t}" for t in limited_texts)

    prompt = _GENERATE_MARKERS_PROMPT.format(reason_name=req.reason_name, texts=texts_block)

    try:
        raw = await _llm_complete(provider, "Ты — специалист по анализу заявок техподдержки.", prompt, temperature=0.1)
        markers_data = _extract_json_from_llm(raw)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"LLM generate error for '{req.reason_name}': {e}")
        raise HTTPException(status_code=502, detail=f"Ошибка LLM: {e}")

    reason_id = _slugify(req.reason_name)

    return GeneratedReason(
        id=reason_id,
        name=req.reason_name,
        markers={
            "verbs": markers_data.get("verbs", []),
            "nouns": markers_data.get("nouns", []),
            "numeric_tags": markers_data.get("numeric_tags", []),
            "phrase_masks": markers_data.get("phrase_masks", []),
        },
        ticket_count=len(texts),
    )


@router.post("/generate-batch", response_model=GenerateBatchResponse, dependencies=[_editor])
async def generate_batch(req: GenerateBatchRequest):
    """Batch-генерация маркеров для нескольких причин."""
    groups = _uploaded_tables.get(req.session_id)
    if groups is None:
        raise HTTPException(status_code=404, detail="Сессия не найдена. Загрузите таблицу заново.")

    provider = _normalize_provider(req.llm_provider)
    results: list[GeneratedReason] = []
    errors: list[dict] = []

    for name in req.reason_names:
        texts = groups.get(name)
        if texts is None:
            errors.append({"reason_name": name, "error": "Не найдена в загруженных данных"})
            continue

        limited_texts = [t[:500] for t in texts[:150]]
        texts_block = "\n".join(f"- {t}" for t in limited_texts)
        prompt = _GENERATE_MARKERS_PROMPT.format(reason_name=name, texts=texts_block)

        try:
            raw = await _llm_complete(
                provider, "Ты — специалист по анализу заявок техподдержки.", prompt, temperature=0.1
            )
            markers_data = _extract_json_from_llm(raw)
            results.append(
                GeneratedReason(
                    id=_slugify(name),
                    name=name,
                    markers={
                        "verbs": markers_data.get("verbs", []),
                        "nouns": markers_data.get("nouns", []),
                        "numeric_tags": markers_data.get("numeric_tags", []),
                        "phrase_masks": markers_data.get("phrase_masks", []),
                    },
                    ticket_count=len(texts),
                )
            )
        except Exception as e:
            logger.error(f"Batch generate error for '{name}': {e}")
            errors.append({"reason_name": name, "error": str(e)})

    return GenerateBatchResponse(results=results, errors=errors)


@router.post("/add-reason", dependencies=[_editor])
async def add_reason(req: AddReasonRequest):
    """Добавить сгенерированную причину в систему."""
    try:
        reason = ContactReason(**req.reason)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Невалидная причина: {e}")

    upsert_reason(reason)
    return {"status": "ok", "id": reason.id, "name": reason.name}


@router.post("/export-reason-docx", dependencies=[_any_admin])
async def export_reason_docx(req: ExportReasonRequest):
    """Экспортировать причину в DOCX (формат шаблона)."""
    from app.api.bot_config import _build_reason_docx, _make_export_filename

    try:
        reason = ContactReason(**req.reason)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Невалидная причина: {e}")

    buffer = _build_reason_docx(reason)
    filename = _make_export_filename(reason)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )
